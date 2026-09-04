"""공문 파일에서 본문 텍스트를 뽑아낸다.

지원 형식
  .hwpx  표준 압축 XML  — 외부 라이브러리 없이 처리
  .hwp   한글 바이너리 — olefile 필요
  .pdf                 — pypdf 필요
  .docx                — 외부 라이브러리 없이 처리
  .txt / .md           — 그대로 읽음

라이브러리가 없으면 그 형식만 건너뛰고 나머지는 정상 동작한다.
"""

from __future__ import annotations

import re
import struct
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import hwpx_view

SUPPORTED = {".hwpx", ".hwp", ".pdf", ".docx", ".xlsx", ".xlsm", ".xls", ".txt", ".md"}

# HWP 문단 안에서 8개 WCHAR(16바이트)를 차지하는 제어 문자들
_WIDE_CONTROLS = {1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23}
_HWPTAG_PARA_TEXT = 67


class ExtractError(Exception):
    pass


def extract_text(path: str | Path) -> str:
    return extract_rich(path)[0]


def extract_rich(path: str | Path) -> tuple[str, str]:
    """(평문, 미리보기 html)을 돌려준다. html은 hwpx에서만 나온다."""
    path = Path(path)
    suffix = path.suffix.lower()
    html = ""
    if suffix == ".hwpx":
        try:
            html, text = hwpx_view.render(path)
        except Exception:          # 구조가 예상과 다르면 평평하게라도 읽는다
            html, text = "", _from_hwpx(path)
    elif suffix == ".hwp":
        text = _from_hwp(path)
    elif suffix == ".pdf":
        text = _from_pdf(path)
    elif suffix == ".docx":
        text = _from_docx(path)
    elif suffix in (".xlsx", ".xlsm", ".xls"):
        html, text = _from_excel(path)
    elif suffix in (".txt", ".md"):
        # utf-8-sig 로 읽어야 메모장이 앞에 붙이는 보이지 않는 표식(BOM)이
        # 벗겨진다. 그냥 utf-8 로 읽으면 그 글자가 제목 앞에 딸려 들어가
        # 목록에 이상한 글자로 보인다. 실제 exe 로 확인한 것이다.
        text = path.read_text(encoding="utf-8-sig", errors="ignore")
    else:
        raise ExtractError(f"지원하지 않는 형식입니다: {suffix}")
    return _tidy(text), html


def _tidy(text: str) -> str:
    # 짝이 깨진 유니코드 문자는 데이터베이스에 넣을 수 없다
    text = text.encode("utf-8", "replace").decode("utf-8")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\u00a0]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(lines).strip()


# ---------------------------------------------------------------- hwpx

def _from_hwpx(path: Path) -> str:
    chunks: list[str] = []
    with zipfile.ZipFile(path) as archive:
        names = [n for n in archive.namelist() if re.fullmatch(r"Contents/section\d+\.xml", n)]
        names.sort(key=lambda n: int(re.search(r"(\d+)", n.split("/")[-1]).group(1)))
        if not names:
            names = [n for n in archive.namelist() if n.endswith(".xml") and "section" in n.lower()]
        for name in names:
            try:
                root = ET.fromstring(archive.read(name).decode("utf-8", "ignore"))
            except ET.ParseError:
                continue
            chunks.append(_walk_hwpx(root))
    if not chunks:
        raise ExtractError("hwpx 안에서 본문 XML을 찾지 못했습니다.")
    return "\n".join(chunks)


def _walk_hwpx(root) -> str:
    out: list[str] = []
    for element in root.iter():
        tag = element.tag.rsplit("}", 1)[-1]
        if tag == "p":
            out.append("\n")
        elif tag == "t" and element.text:
            out.append(element.text)
        elif tag in ("lineBreak", "tab"):
            out.append(" ")
    return "".join(out)


# ----------------------------------------------------------------- hwp

def _from_hwp(path: Path) -> str:
    try:
        import olefile
    except ImportError as exc:  # pragma: no cover
        raise ExtractError("hwp를 읽으려면 olefile이 필요합니다. pip install olefile") from exc
    import zlib

    ole = olefile.OleFileIO(str(path))
    try:
        header = ole.openstream("FileHeader").read()
        compressed = bool(header[36] & 0x01)
        encrypted = bool(header[36] & 0x02)
        if encrypted:
            raise ExtractError("암호가 걸린 hwp 파일입니다.")

        sections = [e for e in ole.listdir() if len(e) > 1 and e[0] == "BodyText"]
        sections.sort(key=lambda e: int(re.sub(r"\D", "", e[1]) or 0))
        if not sections:
            raise ExtractError("hwp 본문 스트림이 없습니다.")

        out: list[str] = []
        for entry in sections:
            data = ole.openstream(entry).read()
            if compressed:
                try:
                    data = zlib.decompress(data, -15)
                except zlib.error:
                    continue
            out.append(_parse_hwp_records(data))
        return "\n".join(out)
    finally:
        ole.close()


def _parse_hwp_records(data: bytes) -> str:
    out: list[str] = []
    cursor, total = 0, len(data)
    while cursor + 4 <= total:
        (raw_header,) = struct.unpack_from("<I", data, cursor)
        cursor += 4
        tag = raw_header & 0x3FF
        size = (raw_header >> 20) & 0xFFF
        if size == 0xFFF:
            if cursor + 4 > total:
                break
            (size,) = struct.unpack_from("<I", data, cursor)
            cursor += 4
        if cursor + size > total:
            break
        if tag == _HWPTAG_PARA_TEXT:
            out.append(_decode_hwp_paragraph(data[cursor:cursor + size]))
        cursor += size
    return "\n".join(out)


_NEWLINE_LE = "\n".encode("utf-16-le")
_SPACE_LE = " ".encode("utf-16-le")


def _decode_hwp_paragraph(raw: bytes) -> str:
    """본문 바이트를 UTF-16 그대로 모았다가 한 번에 해독한다.

    코드 단위를 하나씩 chr() 로 바꾸면 이모지처럼 두 칸을 쓰는 글자가
    반쪽만 남아 저장할 때 깨진다. 바이트로 모아 두면 그런 일이 없다.
    """
    out = bytearray()
    i, size = 0, len(raw) - 1
    while i < size:
        (code,) = struct.unpack_from("<H", raw, i)
        if code in (0, 10, 13):
            out += _NEWLINE_LE
            i += 2
        elif code in _WIDE_CONTROLS:
            out += _SPACE_LE
            i += 16
        elif code < 32:
            i += 2
        else:
            out += raw[i:i + 2]
            i += 2
    return out.decode("utf-16-le", errors="replace")


# ----------------------------------------------------------------- pdf

def _from_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise ExtractError("pdf를 읽으려면 pypdf가 필요합니다. pip install pypdf") from exc
    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "") for page in reader.pages[:20]]
    text = "\n".join(pages)
    if not text.strip():
        raise ExtractError("스캔본으로 보입니다. 글자가 들어 있지 않습니다.")
    return text


# --------------------------------------------------------------- excel

def _from_excel(path: Path) -> tuple[str, str]:
    """제출 서식은 시트마다 안내하는 내용이 다르다. 시트별로 나눠 읽는다."""
    sheets = _read_xls(path) if path.suffix.lower() == ".xls" else _read_xlsx(path)
    if not sheets:
        raise ExtractError("읽을 수 있는 시트가 없습니다.")

    html_parts: list[str] = []
    text_parts: list[str] = []
    for name, rows in sheets:
        text_parts.append(f"[시트: {name}]")
        html_parts.append(f'<p class="hx-sheet">{_esc(name)}</p>')
        if not rows:
            text_parts.append("(빈 시트)")
            html_parts.append('<p class="hx-p hx-obj">빈 시트</p>')
            continue
        text_parts += [" | ".join(cell for cell in row if cell) for row in rows]
        html_parts.append(_rows_to_html(rows))
    return "".join(html_parts), "\n".join(text_parts)


def _read_xlsx(path: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExtractError("엑셀을 읽으려면 openpyxl이 필요합니다. pip install openpyxl") from exc
    book = load_workbook(str(path), read_only=True, data_only=True)
    try:
        out = []
        for sheet in book.worksheets:
            if sheet.sheet_state != "visible":
                continue
            out.append((sheet.title, _trim(_cells(sheet.iter_rows(values_only=True)))))
        return out
    finally:
        book.close()


def _read_xls(path: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ExtractError("구형 xls를 읽으려면 xlrd가 필요합니다. pip install xlrd") from exc
    book = xlrd.open_workbook(str(path))
    out = []
    for sheet in book.sheets():
        rows = ([sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows))
        out.append((sheet.name, _trim(_cells(rows))))
    return out


def _cells(rows) -> list[list[str]]:
    out: list[list[str]] = []
    for row in rows:
        values = ["" if value is None else str(value).strip() for value in row]
        out.append(values)
        if len(out) >= 200:              # 서식 미리보기에 이 정도면 충분하다
            break
    return out


def _trim(rows: list[list[str]]) -> list[list[str]]:
    """바깥쪽 빈 줄과 빈 칸을 걷어낸다."""
    while rows and not any(cell for cell in rows[-1]):
        rows.pop()
    while rows and not any(cell for cell in rows[0]):
        rows.pop(0)
    if not rows:
        return []
    width = max((max((i + 1 for i, c in enumerate(r) if c), default=0) for r in rows), default=0)
    return [row[:width] for row in rows]


def _rows_to_html(rows: list[list[str]]) -> str:
    out = ['<table class="hx-table">']
    for index, row in enumerate(rows[:60]):
        tag = "th" if index == 0 else "td"
        out.append("<tr>" + "".join(f"<{tag}>{_esc(cell)}</{tag}>" for cell in row) + "</tr>")
    out.append("</table>")
    if len(rows) > 60:
        out.append(f'<p class="hx-p hx-obj">아래로 {len(rows) - 60}줄 더 있습니다</p>')
    return "".join(out)


def _esc(text: str) -> str:
    return (text.replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))


# ---------------------------------------------------------------- docx

def _from_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml").decode("utf-8", "ignore")
    root = ET.fromstring(xml)
    out: list[str] = []
    for element in root.iter():
        tag = element.tag.rsplit("}", 1)[-1]
        if tag == "p":
            out.append("\n")
        elif tag == "t" and element.text:
            out.append(element.text)
    return "".join(out)
